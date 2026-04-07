"""Funciones para seleccionar archivos, leer Excel y exportar rangos como imágenes."""

import os
import time
import tkinter as tk
from tkinter import filedialog

import openpyxl
import win32com.client as win32
from openpyxl.utils import get_column_letter
from PIL import ImageGrab

import state
from config import SSO_MARCADOR_TABLA

# Crea una ventana raíz oculta para usar los diálogos de selección sin mostrar una ventana principal.
def _crear_root_oculto():
    root = tk.Tk()
    root.attributes('-topmost', True)
    root.withdraw()
    return root

# Abre un selector para escoger un archivo.
def seleccionar_archivo(mensaje):
    print(f"\nSelecciona {mensaje}:")
    root = _crear_root_oculto()
    try:
        archivo = filedialog.askopenfilename(title=f"Selecciona {mensaje}")
    finally:
        root.destroy()
    if not archivo:
        state.errores.append(f"[ERROR] No se seleccionó {mensaje}")
    return archivo

# Abre un selector para escoger una carpeta.
def seleccionar_carpeta():
    root = _crear_root_oculto()
    try:
        carpeta = filedialog.askdirectory(title="Selecciona carpeta para guardar el informe")
    finally:
        root.destroy()
    return carpeta

# Obtiene o crea una instancia reutilizable de Excel por COM.
def _obtener_excel_app():
    if state._excel_app is None:
        state._excel_app = win32.gencache.EnsureDispatch("Excel.Application")
        state._excel_app.Visible = False
    return state._excel_app

# Cierra todos los libros abiertos y finaliza la instancia de Excel.
def cerrar_excels():
    if state._excel_app is not None:
        try:
            for wb in list(state._excel_app.Workbooks):
                wb.Close(False)
            state._excel_app.Quit()
        except Exception:
            pass
        state._excel_app = None
        state._workbooks_abiertos = {}

# Espera y recupera la imagen copiada al portapapeles desde Excel.
def _clipboard_imagen_post_copy():
    """Espera y reintenta leer el portapapeles (CopyPicture a veces tarda)."""
    for espera in (1.2, 0.6, 0.6):
        time.sleep(espera)
        img = ImageGrab.grabclipboard()
        if img is not None:
            return img
    return None

# Exporta un rango de Excel como imagen.
def exportar_imagen_excel(ruta_excel, hoja, rango, nombre_imagen):
    return exportar_imagen_excel_rangos(ruta_excel, hoja, [rango], nombre_imagen)

# Exporta uno o varios rangos de Excel como una sola imagen.
def exportar_imagen_excel_rangos(ruta_excel, hoja, lista_rangos, nombre_imagen):
    """
    Copia una o más áreas de Excel como imagen. Si hay varios rangos, usa Union
    (útil para repetir fila de encabezado + cuerpo).
    """
    if not lista_rangos:
        state.errores.append(f"[ERROR] Sin rangos para exportar ({nombre_imagen})")
        return os.path.join(r"C:\\Temp", nombre_imagen)

    carpeta_temp = r"C:\\Temp"
    if not os.path.exists(carpeta_temp):
        os.makedirs(carpeta_temp)

    imagen_salida = os.path.join(carpeta_temp, nombre_imagen)
    rango_desc = ",".join(lista_rangos)

    try:
        excel = _obtener_excel_app()
        wb = state._workbooks_abiertos.get(ruta_excel)
        if wb is None:
            wb = excel.Workbooks.Open(ruta_excel, UpdateLinks=0)
            state._workbooks_abiertos[ruta_excel] = wb

        ws = wb.Worksheets(hoja)
        rng = ws.Range(lista_rangos[0])
        for addr in lista_rangos[1:]:
            rng = excel.Union(rng, ws.Range(addr))
        rng.CopyPicture(Appearance=1, Format=2)

        img = _clipboard_imagen_post_copy()
        if img:
            img.save(imagen_salida, "PNG")
        else:
            state.errores.append(f"[ERROR] No se pudo obtener imagen desde el portapapeles ({hoja} {rango_desc})")

    except Exception as e:
        state.errores.append(f"[ERROR] Falló exportación de imagen {hoja} {rango_desc}: {e}")

    return imagen_salida

# Implementa una parte específica de la lógica del informe.
def _fila_tiene_contenido_util(row_vals):
    """True si la fila no es solo vacíos y ceros (para recortar relleno al final)."""
    for v in row_vals:
        if v is None:
            continue
        if isinstance(v, (int, float)):
            if v != 0:
                return True
            continue
        if isinstance(v, str):
            s = v.strip()
            if s == "":
                continue
            try:
                if float(s.replace(",", ".")) != 0:
                    return True
            except ValueError:
                return True
            continue
        return True
    return False

# Implementa una parte específica de la lógica del informe.
def _ultima_fila_con_datos_en_rango_com(ws_com, min_col, min_row, max_col, max_row):
    """Usa la hoja ya abierta en Excel (COM) para no bloquear el archivo con otro lector."""
    for r in range(max_row, min_row - 1, -1):
        vals = [ws_com.Cells(r, c).Value for c in range(min_col, max_col + 1)]
        if _fila_tiene_contenido_util(vals):
            return r
    return min_row

# Implementa una parte específica de la lógica del informe.
def _columna_izquierda_tabla_sso(ws_com, fila_encabezado, max_col_limit=40):
    """Primera columna con texto en la fila de encabezado (por si hay celdas combinadas)."""
    for c in range(1, max_col_limit + 1):
        v = ws_com.Cells(fila_encabezado, c).Value
        if v is not None and str(v).strip() != "":
            return c
    return 1

# Implementa una parte específica de la lógica del informe.
def _ultima_columna_cabecera_sso(ws_com, fila_encabezado, max_col_limit=40):
    """Última columna con texto en la fila del encabezado de la tabla."""
    last = 1
    for c in range(1, max_col_limit + 1):
        v = ws_com.Cells(fila_encabezado, c).Value
        if v is not None and str(v).strip() != "":
            last = c
    return max(last, 1)

# Implementa una parte específica de la lógica del informe.
def _filas_encabezado_tablas_sso(ws_com, marcador=SSO_MARCADOR_TABLA):
    """
    Filas donde empieza cada tabla: celdas A–C de la fila (cabecera suele estar en A o
    en B si hay celdas combinadas). No se busca en columnas de datos para evitar falsos positivos.
    """
    used = ws_com.UsedRange
    max_r = used.Row + used.Rows.Count - 1
    filas = []
    for r in range(1, max_r + 1):
        for c in range(1, 4):
            v = ws_com.Cells(r, c).Value
            if v is None:
                continue
            if marcador in str(v).strip().lower():
                filas.append(r)
                break
    return filas

# Implementa una parte específica de la lógica del informe.
def _rangos_tablas_sso_backup_dinamico(ws_com):
    """
    Una tabla = desde la fila con 'Id del incidente' hasta la fila anterior a la siguiente tabla
    (o fin de datos). Un solo rango contiguo por tabla: siempre incluye encabezados y recorta
    filas vacías/cero al final.
    """
    filas_h = _filas_encabezado_tablas_sso(ws_com)
    if not filas_h:
        state.errores.append(
            "[ERROR] SSO: no se encontró ninguna tabla con encabezado tipo 'Id del incidente'."
        )
        return []

    used = ws_com.UsedRange
    sheet_max_r = used.Row + used.Rows.Count - 1
    rangos = []

    for idx, h_row in enumerate(filas_h):
        next_h = filas_h[idx + 1] if idx + 1 < len(filas_h) else None
        max_row_bloque = (next_h - 1) if next_h else sheet_max_r
        if max_row_bloque < h_row:
            max_row_bloque = h_row

        min_c = _columna_izquierda_tabla_sso(ws_com, h_row)
        last_c = _ultima_columna_cabecera_sso(ws_com, h_row)
        if last_c < min_c:
            last_c = min_c
        last_row = _ultima_fila_con_datos_en_rango_com(ws_com, min_c, h_row, last_c, max_row_bloque)

        rango_a1 = (
            f"{get_column_letter(min_c)}{h_row}:"
            f"{get_column_letter(last_c)}{last_row}"
        )
        rangos.append(rango_a1)

    return rangos

# Extrae el texto resumen desde el Excel base.
def extraer_resumen_excel(ruta_excel):
    try:
        wb = openpyxl.load_workbook(ruta_excel, data_only=True)
        sheet = wb["Grupo Minero FCAB PLAN"]
        resumen = []
        for row in range(38, 44):
            val = sheet[f"B{row}"].value
            if val:
                resumen.append(str(val))
        return "\n".join(resumen)
    except Exception as e:
        state.errores.append(f"[ERROR] No se pudo extraer resumen del Excel madre: {e}")
        return "Resumen no disponible."
